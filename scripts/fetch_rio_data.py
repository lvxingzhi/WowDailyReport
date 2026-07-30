"""
Raider.io 数据自动抓取脚本

从 raider.io API 抓取大秘境数据，自动追加到 Excel「每日数据」Sheet。
全部数据自动填充，无需手动补录。

跨赛季使用：只需修改下方 CONFIG 中的 SEASON_SLUG、EXPANSION_ID 和 EXCEL_FILE。

@author ext.ahs.lvxingz1
"""
import json
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import openpyxl

# ╔══════════════════════════════════════════════════════════════════╗
# ║                        📋 配置区（跨赛季改这里）                  ║
# ╚══════════════════════════════════════════════════════════════════╝

# --- 赛季配置 ---
SEASON_SLUG = "season-mn-1"     # 赛季 slug，如 season-mn-1, season-mn-2, season-tww-1
EXPANSION_ID = 11               # 资料片 ID：11=Midnight, 10=TheWarWithin
EXCEL_FILE = "大秘境分数行情_至暗之夜第一赛季.xlsx"  # Excel 文件名

# 代理设置（留空则不使用代理）
PROXY = ""

# --- 40 专精映射：Excel 列号（1-based）→ API class/spec ---
# 列号对应：Q=17, R=18, ... BF=56
SPEC_MAP = [
    # (列号, API class_slug, API spec_slug, Excel 中文名)
    (17, "death-knight",  "blood",         "T血DK"),
    (18, "death-knight",  "frost",         "D冰DK"),
    (19, "death-knight",  "unholy",        "D邪DK"),
    (20, "demon-hunter",  "devourer",      "D噬灭"),
    (21, "demon-hunter",  "havoc",         "D浩劫"),
    (22, "demon-hunter",  "vengeance",     "T复仇"),
    (23, "druid",         "balance",       "D鸟德"),
    (24, "druid",         "feral",         "D野德"),
    (25, "druid",         "guardian",      "T熊T"),
    (26, "druid",         "restoration",   "H奶德"),
    (27, "evoker",        "augmentation",  "D增辉"),
    (28, "evoker",        "devastation",   "D湮灭"),
    (29, "evoker",        "preservation",  "H恩护"),
    (30, "hunter",        "beast-mastery", "D兽王"),
    (31, "hunter",        "marksmanship",  "D射击"),
    (32, "hunter",        "survival",      "D生存"),
    (33, "mage",          "arcane",        "D奥法"),
    (34, "mage",          "fire",          "D火法"),
    (35, "mage",          "frost",         "D冰法"),
    (36, "monk",          "brewmaster",    "T酒仙"),
    (37, "monk",          "mistweaver",    "H奶僧"),
    (38, "monk",          "windwalker",    "D踏风"),
    (39, "paladin",       "holy",          "H奶骑"),
    (40, "paladin",       "protection",    "T防骑"),
    (41, "paladin",       "retribution",   "D惩戒"),
    (42, "priest",        "discipline",    "H戒律"),
    (43, "priest",        "holy",          "H神牧"),
    (44, "priest",        "shadow",        "D暗牧"),
    (45, "rogue",         "assassination", "D刺杀"),
    (46, "rogue",         "outlaw",        "D狂徒"),
    (47, "rogue",         "subtlety",      "D敏锐"),
    (48, "shaman",        "elemental",     "D元素"),
    (49, "shaman",        "enhancement",   "D增强"),
    (50, "shaman",        "restoration",   "H奶萨"),
    (51, "warlock",       "affliction",    "D痛苦"),
    (52, "warlock",       "demonology",    "D恶魔"),
    (53, "warlock",       "destruction",   "D毁灭"),
    (54, "warrior",       "arms",          "D武器"),
    (55, "warrior",       "fury",          "D狂暴"),
    (56, "warrior",       "protection",    "T防战"),
]

# --- 8 副本映射：Excel 列号 → API dungeon slug ---
DUNGEON_MAP = [
    # (列号 1-based, API slug, 中文名)
    (9,  "seat-of-the-triumvirate", "执政团之座"),
    (10, "nexuspoint-xenas",        "节点希纳斯"),
    (11, "algethar-academy",        "艾杰斯亚学院"),
    (12, "pit-of-saron",            "萨隆矿坑"),
    (13, "maisara-caverns",         "迈萨拉洞窟"),
    (14, "skyreach",                "通天峰"),
    (15, "windrunner-spire",        "风行者之塔"),
    (16, "magisters-terrace",       "魔导师平台"),
]

# --- 专精简称映射（API spec名 → 中文简称，用于 H 列队伍名）---
SPEC_ABBR = {
    "Blood": "T血DK", "Frost": "D冰DK", "Unholy": "D邪DK",
    "Devourer": "D噬灭", "Havoc": "D浩劫", "Vengeance": "T复仇",
    "Balance": "D鸟德", "Feral": "D野德", "Guardian": "T熊T", "Restoration": "H奶德",
    "Augmentation": "D增辉", "Devastation": "D湮灭", "Preservation": "H恩护",
    "Beast Mastery": "D兽王", "Marksmanship": "D射击", "Survival": "D生存",
    "Arcane": "D奥法", "Fire": "D火法",
    "Brewmaster": "T酒仙", "Mistweaver": "H奶僧", "Windwalker": "D踏风",
    "Holy": "H神牧", "Protection": "T防骑", "Retribution": "D惩戒",
    "Discipline": "H戒律", "Shadow": "D暗牧",
    "Assassination": "D刺杀", "Outlaw": "D狂徒", "Subtlety": "D敏锐",
    "Elemental": "D元素", "Enhancement": "D增强",
    "Affliction": "D痛苦", "Demonology": "D恶魔", "Destruction": "D毁灭",
    "Arms": "D武器", "Fury": "D狂暴",
}
# 处理重名：Holy 在 Paladin 是 H奶骑，在 Priest 是 H神牧
# 以及 Restoration 在 Druid 是 H奶德，在 Shaman 是 H奶萨
# Frost 在 DK 是 D冰DK，在 Mage 是 D冰法
# 用 (class, spec) 元组覆盖
SPEC_ABBR_OVERRIDE = {
    ("Paladin", "Holy"): "H奶骑",
    ("Shaman", "Restoration"): "H奶萨",
    ("Druid", "Restoration"): "H奶德",
    ("Mage", "Frost"): "D冰法",
    ("Death Knight", "Frost"): "D冰DK",
    ("Priest", "Holy"): "H神牧",
    ("Warrior", "Protection"): "T防战",
    ("Paladin", "Protection"): "T防骑",
}


# ╔══════════════════════════════════════════════════════════════════╗
# ║                          🔧 工具函数                             ║
# ╚══════════════════════════════════════════════════════════════════╝

def api_get(url, retries=3, delay=1.0):
    """调用 raider.io API，自动重试。"""
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
                "Accept": "application/json",
            })
            if PROXY:
                req.set_proxy(PROXY, "http")
            resp = urllib.request.urlopen(req, timeout=20)
            return json.loads(resp.read())
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                raise


def spec_abbr(class_name, spec_name):
    """返回中文专精简称，如 'Guardian' + 'Druid' → 'T熊T'"""
    key = (class_name, spec_name)
    if key in SPEC_ABBR_OVERRIDE:
        return SPEC_ABBR_OVERRIDE[key]
    return SPEC_ABBR.get(spec_name, spec_name)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       📊 数据抓取函数                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def _fetch_single_spec(col, class_slug, spec_slug, cn_name):
    """抓取单个专精 #1，供线程池调用。"""
    url = (f"https://raider.io/api/mythic-plus/rankings/specs"
           f"?season={SEASON_SLUG}&region=world"
           f"&class={class_slug}&spec={spec_slug}&page=0")
    try:
        data = api_get(url, retries=2, delay=0.5)
        ranked = data["rankings"]["rankedCharacters"]
        if ranked:
            top = ranked[0]
            char = top["character"]
            return col, {
                "score": top["score"],
                "name": char["name"],
                "class": char["class"]["name"],
                "spec": char["spec"]["name"],
                "run_id": (top.get("runs") or [{}])[0].get("keystoneRunId"),
            }
        return col, {"score": 0, "name": "", "class": "", "spec": "", "run_id": None}
    except Exception as e:
        print(f"  ⚠️ 专精 {cn_name} ({class_slug}/{spec_slug}) 抓取失败: {e}")
        return col, {"score": 0, "name": "", "class": "", "spec": "", "run_id": None}


def fetch_spec_scores():
    """并发抓取 40 专精全球 #1 分数。返回 {col: {...}}"""
    results = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {
            executor.submit(_fetch_single_spec, col, cs, ss, cn): (col, cn)
            for col, cs, ss, cn in SPEC_MAP
        }
        for future in as_completed(futures):
            col, info = future.result()
            results[col] = info
            cn_name = futures[future][1]
            print(f"  ✅ {cn_name}: {info['score']:.1f}")
    return results


def _fetch_single_dungeon(col, slug, cn_name):
    """抓取单个副本最高层数，供线程池调用。"""
    url = (f"https://raider.io/api/v1/mythic-plus/runs"
           f"?season={SEASON_SLUG}&region=world&dungeon={slug}&page=0")
    try:
        data = api_get(url, retries=2, delay=0.5)
        if data.get("rankings"):
            return col, data["rankings"][0]["run"]["mythic_level"]
        return col, 0
    except Exception as e:
        print(f"  ⚠️ 副本 {cn_name} ({slug}) 抓取失败: {e}")
        return col, 0


def fetch_dungeon_levels():
    """并发抓取 8 副本全球最高限时层数。返回 {col: level}"""
    results = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {
            executor.submit(_fetch_single_dungeon, col, slug, cn): (col, cn)
            for col, slug, cn in DUNGEON_MAP
        }
        for future in as_completed(futures):
            col, level = future.result()
            results[col] = level
            cn_name = futures[future][1]
            print(f"  ✅ {cn_name}: {level}")
    return results


def fetch_season_cutoffs():
    """抓取中国区 0.1% / 1% 分数线 + 人数。返回 (top01pct, top1pct, total, pop01, pop1)"""
    url = (f"https://raider.io/api/v1/mythic-plus/season-cutoffs"
           f"?season={SEASON_SLUG}&region=cn")
    data = api_get(url)
    cutoffs = data["cutoffs"]
    top01 = cutoffs["p999"]["all"]["quantileMinValue"]
    top1 = cutoffs["p990"]["all"]["quantileMinValue"]
    total = cutoffs["p999"]["all"]["totalPopulationCount"]
    pop01 = cutoffs["p999"]["all"]["quantilePopulationCount"]
    pop1 = cutoffs["p990"]["all"]["quantilePopulationCount"]
    return round(top01, 2), round(top1, 2), total, pop01, pop1


def fetch_exact_rank_score(total, pct):
    """
    精确查排行榜第 total*pct 名分数。
    返回 (score, rank) 或 (None, 0)
    """
    rank = int(total * pct)
    page = (rank - 1) // 40
    idx = (rank - 1) % 40
    url = (f"https://raider.io/api/mythic-plus/rankings/characters"
           f"?season={SEASON_SLUG}&region=cn&class=all&role=all&page={page}&pageSize=40")
    data = api_get(url)
    ranked = data["rankings"]["rankedCharacters"]
    score = ranked[idx]["score"]
    return round(score, 2), rank


def fetch_top_characters():
    """
    获取全球最高分及所有并列最高分角色。
    调全局排行榜 API（按分数降序），逐页收集直到分数低于最高分。
    返回 (highest_score, chars_str)
      chars_str: "角色名-专精, 角色名-专精, ..."
    """
    print("  🔍 正在查找全球最高分角色...")

    url = (f"https://raider.io/api/mythic-plus/rankings/characters"
           f"?season={SEASON_SLUG}&region=world&class=all&role=all"
           f"&page=0&pageSize=40")
    try:
        data = api_get(url)
        ranked = data["rankings"]["rankedCharacters"]
        if not ranked:
            print("  ⚠️ 未能获取最高分")
            return 0, ""
        best_score = ranked[0]["score"]
    except Exception as e:
        print(f"  ⚠️ 获取全球最高分失败: {e}")
        return 0, ""

    # 收集所有并列最高分的角色（跨页）
    top_chars = []
    page = 0
    while True:
        if page > 0:
            url = (f"https://raider.io/api/mythic-plus/rankings/characters"
                   f"?season={SEASON_SLUG}&region=world&class=all&role=all"
                   f"&page={page}&pageSize=40")
            data = api_get(url)
            ranked = data["rankings"]["rankedCharacters"]

        for entry in ranked:
            if entry["score"] < best_score:
                # 分数开始下降，停止收集
                break
            char = entry["character"]
            abbr = spec_abbr(char["class"]["name"], char["spec"]["name"])
            top_chars.append(f"{char['name']}-{abbr}")
        else:
            # 本页全部都是最高分，继续翻下一页
            page += 1
            time.sleep(0.15)
            continue
        # 遇到分数下降，退出
        break

    chars_str = ",".join(top_chars)
    print(f"  最高分: {best_score:.1f}（{len(top_chars)} 人）: {chars_str}")
    return round(float(best_score), 1), chars_str





# ╔══════════════════════════════════════════════════════════════════╗
# ║                       📝 Excel 写入                             ║
# ╚══════════════════════════════════════════════════════════════════╝

def append_to_excel(excel_path, row_data):
    """
    向 Excel「每日数据」Sheet 追加一行。

    参数:
        excel_path: Excel 文件路径
        row_data: dict，key 为 1-based 列号，value 为单元格值
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["每日数据"]

    next_row = ws.max_row + 1
    for col, value in row_data.items():
        ws.cell(row=next_row, column=col, value=value)

    wb.save(excel_path)
    wb.close()
    print(f"✅ 已追加到 {excel_path} 第 {next_row} 行")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                          🚀 主流程                              ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    project_root = Path(__file__).resolve().parent.parent
    excel_path = project_root / EXCEL_FILE

    if not excel_path.exists():
        print(f"❌ Excel 文件不存在: {excel_path}")
        sys.exit(1)

    today_str = datetime.now().strftime("%Y%m%d")
    print(f"🚀 开始抓取数据... (日期={today_str}, 赛季={SEASON_SLUG})")
    print()

    # 1. 分数线（中国区）
    print("📊 分数线 (CN)...")
    top01pct, top1pct, total_pop, pop01, pop1 = fetch_season_cutoffs()
    print(f"  0.1%: {top01pct}, 1%: {top1pct}  (总人数={total_pop}, 0.1%人数={pop01}, 1%人数={pop1})")

    # 2.5 精确分数线（查排行榜第N名）
    def try_exact(pct, label):
        try:
            score, rank = fetch_exact_rank_score(total_pop, pct)
            print(f"  {label}: {score} (排名={rank})")
            return score
        except Exception as e:
            print(f"  ⚠️ {label}: 抓取失败 ({e})")
            return None

    top09pct = try_exact(0.009, "0.9%")
    top009pct = try_exact(0.0009, "0.09%")
    print()

    # 3. 副本层数（全球）
    print("🗺️ 副本限时层数 (World)...")
    dungeon_levels = fetch_dungeon_levels()
    for col, (slug, cn_name) in [(c, (s, n)) for c, s, n in DUNGEON_MAP]:
        print(f"  {cn_name}: {dungeon_levels.get(col, '?')}")
    print()

    # 4. 专精分数（全球）+ 顺便找最高分队伍
    print("⚔️ 专精最高分 (World)...")
    spec_scores = fetch_spec_scores()
    for col, _cs, _ss, cn_name in SPEC_MAP:
        info = spec_scores.get(col, {})
        print(f"  {cn_name}: {info.get('score', 0):.1f}")
    print()

    # 5. 最高分+队伍（复用 spec_scores，不再重复请求 API）
    print("👑 最高分队伍...")
    highest_score, top_chars = fetch_top_characters()
    print()

    # 6. 组装 Excel 行数据
    # 列号映射：
    # A=1(日期), B=2(赛季), C=3(赛季周), D=4(剩余周),
    # E=5(最高分), F=6(0.1%), G=7(1%), H=8(最高分角色),
    # I-P=9~16(副本), Q-BD=17~56(专精),
    # BE=57(0.9%), BF=58(0.09%), BG=59(总人数)
    # 注：B~D 列（赛季名/周数/剩余周）由用户自行填写，脚本不自动计算

    row_data = {
        1: int(today_str),
        5: highest_score,
        6: top01pct,
        7: top1pct,
        8: top_chars,
    }

    # 副本层数
    for col, level in dungeon_levels.items():
        row_data[col] = level

    # 专精分数
    for col, info in spec_scores.items():
        row_data[col] = round(info["score"], 1)

    # 精确分数线（列 BE=57/BF=58，在专精之后）
    if top09pct is not None:
        row_data[57] = top09pct
    if top009pct is not None:
        row_data[58] = top009pct

    # 总人数（列 BG=59）
    row_data[59] = total_pop
    # 0.1%人数（列 BH=60）
    row_data[60] = pop01
    # 1%人数（列 BI=61）
    row_data[61] = pop1

    # 7. 写入 Excel
    append_to_excel(excel_path, row_data)

    # 8. 完成
    print()
    print("🎉 完成！全部数据已自动填充。")


if __name__ == "__main__":
    main()
