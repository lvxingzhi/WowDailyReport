"""
大秘境分数行情 Excel 数据解析脚本

职责：读取 Excel 每日数据 Sheet，解析为 JSON 格式，
      生成 current.json 和按日期归档的副本。

@author ext.ahs.lvxingz1
"""
import json
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl


# 项目根目录（scripts 的上级目录）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = PROJECT_ROOT / "大秘境分数行情_至暗之夜第一赛季.xlsx"
DATA_DIR = PROJECT_ROOT / "data"
ARCHIVE_DIR = DATA_DIR / "archive"
CURRENT_FILE = DATA_DIR / "current.json"

# 强势阵容 Sheet 名称
NATIONAL_TEAM_SHEET = "强势阵容"

# 列索引常量
COL_DATE = 0       # 日期
COL_SEASON = 1     # 赛季
COL_WEEK = 2       # 赛季周
COL_REMAINING = 3  # 距离结束(周)
COL_HIGHEST_SCORE = 4   # 最高分数
COL_TOP01_PCT = 5      # 0.1%分数
COL_TOP1_PCT = 6       # 1%分数
COL_TEAM = 7           # 最高队伍(角色名-职业,逗号分隔)
DUNGEON_START_COL = 8  # 副本层数起始列（8~15，共8个）
COL_IRON24 = 16        # 最高铁人数
COL_IRON23 = 17        # 次铁人数
SPEC_START_COL = 18    # 专精分数起始列（18~57，共40个）


def parse_national_team(ws) -> list:
    """
    解析「强势阵容」Sheet，每行是一套5人阵容。

    第一行为角色头（坦克, 治疗, 输出, 输出, 输出），
    后续每行对应一套阵容。

    参数:
        ws: openpyxl worksheet 对象（强势阵容 Sheet）

    返回:
        list: [{ "group": "阵容1", "members": [{ "role": "坦克", "class": "T熊T" }, ...] }, ...]
    """
    headers = [cell.value for cell in ws[1]]
    lineups = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        members = []
        for i, class_val in enumerate(row):
            if class_val is not None and str(class_val).strip():
                role = headers[i] if i < len(headers) and headers[i] else "输出"
                members.append({
                    "role": role,
                    "class": str(class_val).strip()
                })
        if members:
            lineups.append({
                "group": f"阵容{row_idx}",
                "members": members
            })
    return lineups


def parse_excel(excel_path: str) -> dict:
    """
    解析 Excel 文件，返回完整的 JSON 数据结构。

    参数:
        excel_path: Excel 文件路径

    返回:
        dict: 包含 meta、daily 和 nationalTeam 的完整数据结构
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["每日数据"]

    # 读取表头
    headers = [cell.value for cell in ws[1]]

    # 获取副本名称列表（第 8~15 列）
    dungeon_names = headers[DUNGEON_START_COL:DUNGEON_START_COL + 8]

    # 获取专精名称列表（第 18~57 列，共 40 个）
    spec_names = headers[SPEC_START_COL:]

    daily_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        # --- 日期解析 ---
        raw_date = row[COL_DATE]
        if isinstance(raw_date, (int, float)):
            date_str = str(int(raw_date))
        elif isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y%m%d")
        else:
            date_str = str(raw_date)

        # --- 队伍解析（格式："角色名-职业,角色名-职业,..."） ---
        team_str = str(row[COL_TEAM] or "").strip().replace("\t", "")
        team_entries = [m.strip() for m in team_str.split(",") if m.strip()]

        team_players = []
        rank1_player = ""
        rank1_class = ""
        for i, entry in enumerate(team_entries):
            if "-" in entry:
                parts = entry.rsplit("-", 1)
                player = parts[0].strip()
                cls = parts[1].strip()
            else:
                player = entry
                cls = ""
            team_players.append({"player": player, "class": cls})
            if i == 0:
                rank1_player = player
                rank1_class = cls

        # --- 最高分数解析 ---
        score_raw = row[COL_HIGHEST_SCORE]
        rank1_score = 0
        if score_raw is not None:
            if isinstance(score_raw, str):
                score_raw = score_raw.strip().replace("\t", "")
            try:
                rank1_score = int(float(score_raw))
            except (ValueError, TypeError):
                rank1_score = 0

        # --- 安全转换函数 ---
        def safe_int(val, default=0):
            if val is None:
                return default
            if isinstance(val, str):
                val = val.strip().replace("\t", "")
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return val  # 保留非数字原始值（如 "?"）

        def safe_float(val):
            if val is None:
                return 0.0
            if isinstance(val, str):
                val = val.strip().replace("\t", "")
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        top01pct = safe_float(row[COL_TOP01_PCT])
        top1pct = safe_float(row[COL_TOP1_PCT])

        # --- 铁钥匙人数 ---
        iron24 = safe_int(row[COL_IRON24])
        iron23 = safe_int(row[COL_IRON23])

        # --- 专精分数（40 个） ---
        faith_specs = []
        for i, name in enumerate(spec_names):
            col_idx = SPEC_START_COL + i
            score = safe_float(row[col_idx] if col_idx < len(row) else None)
            faith_specs.append({
                "name": name if name else f"专精{i + 1}",
                "score": score
            })
        # 按分数从高到低排序
        faith_specs.sort(key=lambda x: x["score"], reverse=True)

        # --- 副本数据 ---
        dungeons = []
        for i, name in enumerate(dungeon_names):
            col_idx = DUNGEON_START_COL + i
            level = row[col_idx] if col_idx < len(row) else None
            dungeons.append({
                "name": name if name else f"副本{i + 1}",
                "level": int(level) if level is not None else 0
            })

        day_data = {
            "date": date_str,
            "season": row[COL_SEASON],
            "seasonWeek": safe_int(row[COL_WEEK]),
            "weeksRemaining": safe_int(row[COL_REMAINING]),
            "rank1": {
                "score": rank1_score,
                "player": rank1_player,
                "class": rank1_class,
                "team": team_players
            },
            "top1Pct": top1pct,
            "top01Pct": top01pct,
            "iron24": iron24,
            "iron23": iron23,
            "faithSpecs": faith_specs,
            "dungeons": dungeons
        }
        daily_list.append(day_data)

    # 解析强势阵容
    national_team = []
    if NATIONAL_TEAM_SHEET in wb.sheetnames:
        national_team = parse_national_team(wb[NATIONAL_TEAM_SHEET])

    wb.close()

    # 构建 meta
    season = daily_list[0]["season"] if daily_list else ""
    latest_date = daily_list[-1]["date"] if daily_list else ""

    result = {
        "meta": {
            "season": season,
            "generatedAt": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "dataCount": len(daily_list),
            "latestDate": latest_date
        },
        "daily": daily_list,
        "nationalTeam": national_team
    }

    return result


def save_json(data: dict, filepath: Path) -> None:
    """
    将数据保存为 JSON 文件，UTF-8 编码，美化格式。

    参数:
        data: 要保存的数据字典
        filepath: 目标文件路径
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    """主入口：解析 Excel 并生成 JSON 文件。"""
    if not EXCEL_FILE.exists():
        print(f"[错误] Excel 文件不存在: {EXCEL_FILE}")
        sys.exit(1)

    print(f"[开始] 解析 Excel: {EXCEL_FILE}")
    data = parse_excel(str(EXCEL_FILE))
    print(f"[完成] 共解析 {data['meta']['dataCount']} 条数据")

    # 生成 current.json
    save_json(data, CURRENT_FILE)
    print(f"[输出] current.json → {CURRENT_FILE}")

    # 按最新日期归档
    latest_date = data["meta"]["latestDate"]
    if latest_date:
        archive_file = ARCHIVE_DIR / f"{latest_date}.json"
        save_json(data, archive_file)
        print(f"[归档] {latest_date}.json → {archive_file}")

    # 支持命令行参数 --archive YYYYMMDD 归档指定日期之前的数据
    if len(sys.argv) >= 3 and sys.argv[1] == "--archive":
        target_date_str = sys.argv[2]
        try:
            target_date = datetime.strptime(target_date_str, "%Y%m%d").date()
            filtered = [d for d in data["daily"] if _parse_date(d["date"]) <= target_date]
            archive_data = {
                "meta": {**data["meta"], "dataCount": len(filtered), "latestDate": target_date_str},
                "daily": filtered
            }
            archive_file = ARCHIVE_DIR / f"{target_date_str}.json"
            save_json(archive_data, archive_file)
            print(f"[归档] 指定日期 {target_date_str}.json → {archive_file}")
        except ValueError:
            print(f"[错误] 日期格式无效: {target_date_str}，应为 YYYYMMDD")
            sys.exit(1)

    print("[完成] 所有操作已完成")


def _parse_date(date_str: str) -> date:
    """将 YYYYMMDD 字符串解析为 date 对象。"""
    return datetime.strptime(date_str, "%Y%m%d").date()


if __name__ == "__main__":
    main()
